#!/usr/bin/env python
"""M6F — census certification: the complete realized-stream clinician audit (Table 9R).

Merges the M6-400 stratified audit (both clinicians, released) with the returned M6F census
block (730 remaining accepted + 102 rejected, both clinicians) into full double clinician labels
for the realized 1,232-row one-answer stream at the frozen threshold tau* = 0.412661, and
computes, EXACTLY as prespecified in refine-logs/analysis_plan_v2_20260717.md:

  4.5  consensus rules: primary = worst-case (either clinician labels error OR unsure);
       also both-confirm; census quantities are finite-population exact proportions
       (CP intervals shown only as superpopulation companions, labeled);
  4.6  Table 9R rows (stage-1 proxy / both / either-worst-case) x columns (whole-stream /1232,
       accepted /1130, moderate-or-worse accepted /1130, decision at fixed alpha = 0.20) with
       the LOCKED three-outcome wording;
  5.2  decision-consequence analysis per policy (defer-none / defer-all / global-only tau* /
       beta*-constrained, which at this frozen operating point blocks deployment since the
       certified paired bound exceeds beta* = 0.10): accepted errors per 1,000, accepted
       moderate-or-worse per 1,000, avoided vs global-only, extra deferrals, deferrals per
       moderate-or-worse avoided, and a clinician-label-based decision-curve proxy
       NB(p_t) = TP/N - FP/N * p_t/(1-p_t) (defer = intervention, moderate-or-worse erroneous
       answer = event). This is a label-based proxy, NOT clinical net benefit.

No new certificates, no multiplicity spending, no released number touched. Harm grades:
max of available clinician grades; flagged rows with no grade (unsure-only) are counted
separately and included in the worst-case-inclusive harm count.

Usage: python M6F_certify_census.py <doctor1_formal.xlsx> <doctor2_formal.xlsx>
Output: M6F_census_results.json (local; gitignored) + printed Table 9R.
"""
import os, sys, csv, json
import numpy as np
from openpyxl import load_workbook
from scipy.stats import beta as sbeta

from _paths import artifact                                # released layout, docs/LAYOUT.md

HERE = os.path.dirname(os.path.abspath(__file__))          # <package>/code
ROOT = os.path.dirname(HERE)                               # <package>
ALPHA = 0.20
BETA_STAR = 0.10
SEV_ORDER = {"": 0, "none": 0, "mild": 1, "moderate": 2, "severe": 3, "death_extreme": 4}
PT_GRID = [0.05, 0.10, 0.15, 0.20, 0.30]


def cp_int(k, n, conf=0.95):
    lo = 0.0 if k == 0 else float(sbeta.ppf((1 - conf) / 2, k, n - k + 1))
    hi = 1.0 if k == n else float(sbeta.ppf(1 - (1 - conf) / 2, k + 1, n - k))
    return [round(lo, 4), round(hi, 4)]


def read_sheet(xlsx):
    ws = load_workbook(xlsx, data_only=True)["标注表"]
    hdr = [c.value for c in ws[1]]
    col = {k: hdr.index(k) for k in ("build_id", "blind_id", "clinician_verdict", "severity")}
    rows, builds = {}, set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        bid = str(r[col["blind_id"]] or "")
        if not bid:
            continue
        builds.add(str(r[col["build_id"]]))
        rows[bid] = (str(r[col["clinician_verdict"]] or "").strip().upper(),
                     str(r[col["severity"]] or "").strip().lower())
    return rows, builds


def main():
    f1, f2 = sys.argv[1], sys.argv[2]
    meta = json.load(open(artifact("M6F_census_meta.json")))
    expect_build = meta["build_id"]

    d1, b1 = read_sheet(f1); d2, b2 = read_sheet(f2)
    assert b1 == {expect_build} and b2 == {expect_build}, ("build_id mismatch", b1, b2)

    key = list(csv.DictReader(open(artifact("M6F_census_key.csv"))))
    formal = [r for r in key if r["block"] == "formal"]
    mains = [r for r in formal if r["kind"] == "main"]
    reps = [r for r in formal if r["kind"] == "repeat"]
    assert len(mains) == 832 and len(reps) == 48, (len(mains), len(reps))
    for r in formal:
        assert r["blind_id"] in d1 and r["blind_id"] in d2, f"missing row {r['blind_id']}"
        assert d1[r["blind_id"]][0] and d2[r["blind_id"]][0], f"unfilled verdict {r['blind_id']}"

    # ---- M6-400 released audit rows (both clinicians) ----
    m6key = {r["blind_id"]: r for r in csv.DictReader(open(artifact("M6_SR_audit_key.csv")))
             if r["blind_id"] != ""}
    m6 = {}
    for a in (1, 2):
        for r in csv.DictReader(open(artifact(f"M6_SR_instrument_annotator{a}_FILLED.csv"))):
            v = r["clinician_verdict"].strip().upper()
            s = (r.get("severity") or "").strip().lower()
            m6.setdefault(r["blind_id"], {})[a] = (v, s)
    assert len(m6) == 400 and all(len(v) == 2 for v in m6.values())

    # ---- unified per-unique-row record set: (accepted, v1,s1, v2,s2, stage1, source) ----
    rows = []
    for bid, per in m6.items():
        k = m6key[bid]
        rows.append(dict(uid=f"m6:{bid}", accepted=1, src=k["source"], stage1=int(k["stage1_H"]),
                         v1=per[1][0], s1=per[1][1], v2=per[2][0], s2=per[2][1]))
    for r in mains:
        bid = r["blind_id"]
        rows.append(dict(uid=f"cf:{bid}", accepted=int(r["accepted"]), src=r["source"],
                         stage1=int(r["stage1_H"]), v1=d1[bid][0], s1=d1[bid][1],
                         v2=d2[bid][0], s2=d2[bid][1]))
    N = len(rows); N_acc = sum(r["accepted"] for r in rows)
    assert (N, N_acc) == (1232, 1130), (N, N_acc)

    def err_both(r):  return r["v1"] == "HALLUCINATED" and r["v2"] == "HALLUCINATED"
    def err_wc(r):    return (r["v1"] in ("HALLUCINATED", "UNSURE")) or (r["v2"] in ("HALLUCINATED", "UNSURE"))
    def sev_max(r):
        return max(SEV_ORDER.get(r["s1"], 0), SEV_ORDER.get(r["s2"], 0))
    def modplus_graded(r, rule):   # flagged by rule AND max available grade >= moderate
        return rule(r) and sev_max(r) >= 2
    def flagged_ungraded(r):
        return err_wc(r) and sev_max(r) == 0

    acc = [r for r in rows if r["accepted"] == 1]
    rej = [r for r in rows if r["accepted"] == 0]

    # ---- Table 9R ----
    def row9(label, flag, sevflag):
        ws_k = sum(flag(r) for r in rows); acc_k = sum(flag(r) for r in acc)
        mod_k = sum(sevflag(r) for r in acc)
        return dict(label=label, whole=f"{ws_k}/1232", whole_prop=round(ws_k / N, 4),
                    accepted=f"{acc_k}/1130", accepted_prop=round(acc_k / N_acc, 4),
                    modplus_accepted=f"{mod_k}/1130", modplus_prop=round(mod_k / N_acc, 4),
                    cp95_accepted_superpop=cp_int(acc_k, N_acc))
    t9 = [row9("stage-1 proxy", lambda r: r["stage1"] == 1, lambda r: False),
          row9("both clinicians", err_both, lambda r: modplus_graded(r, err_both)),
          row9("either / worst case", err_wc, lambda r: modplus_graded(r, err_wc))]
    t9[0]["modplus_accepted"] = "—"; t9[0]["modplus_prop"] = None

    wc_prop = t9[2]["accepted_prop"]; both_prop = t9[1]["accepted_prop"]
    if wc_prop <= ALPHA:
        decision = "the realized accepted set met the fixed clinician-label criterion"
    elif both_prop > ALPHA:
        decision = "the fixed clinician-label criterion was not met"
    else:
        decision = "the conclusion was consensus-sensitive"
    # locked rule ordering: wc<=a -> met; wc>a with both>a -> not met; both<=a<wc -> consensus-sensitive
    if wc_prop > ALPHA and both_prop <= ALPHA:
        decision = "the conclusion was consensus-sensitive"

    # ---- reliability ----
    v1b = np.array([r["v1"] == "HALLUCINATED" for r in rows]); v2b = np.array([r["v2"] == "HALLUCINATED" for r in rows])
    po = float((v1b == v2b).mean())
    p1, p2 = v1b.mean(), v2b.mean()
    pe = p1 * p2 + (1 - p1) * (1 - p2)
    kappa_all = (po - pe) / (1 - pe)
    new_mask = np.array([r["uid"].startswith("cf:") for r in rows])
    v1n, v2n = v1b[new_mask], v2b[new_mask]
    pon = float((v1n == v2n).mean()); p1n, p2n = v1n.mean(), v2n.mean()
    pen = p1n * p2n + (1 - p1n) * (1 - p2n)
    kappa_new = (pon - pen) / (1 - pen)
    n_unsure = sum((r["v1"] == "UNSURE") + (r["v2"] == "UNSURE") for r in rows)
    intra = {}
    main_by_blind = {r["blind_id"]: r for r in mains}
    for a, dd in ((1, d1), (2, d2)):
        agree = tot = 0
        for r in reps:
            main_bid = r["repeat_of_blind"]
            if main_bid in dd and r["blind_id"] in dd:
                tot += 1; agree += dd[main_bid][0] == dd[r["blind_id"]][0]
        intra[f"doctor{a}"] = dict(n=tot, agree=agree, rate=round(agree / tot, 4))

    # ---- harm distribution on worst-case flagged accepted rows ----
    sev_dist = {}
    for r in acc:
        if err_wc(r):
            lab = {0: "ungraded", 1: "mild", 2: "moderate", 3: "severe", 4: "death_extreme"}[sev_max(r)]
            sev_dist[lab] = sev_dist.get(lab, 0) + 1
    n_ungraded_flagged = sum(flagged_ungraded(r) for r in acc)

    # ---- decision-consequence analysis (worst-case rule; both-rule as sensitivity) ----
    def consequences(rule, sevrule):
        ws_err = sum(rule(r) for r in rows)
        acc_err = sum(rule(r) for r in acc); rej_err = sum(rule(r) for r in rej)
        acc_mod = sum(sevrule(r) for r in acc); rej_mod = sum(sevrule(r) for r in rej)
        tot_mod = acc_mod + rej_mod
        pol = {}
        pol["defer-none"] = dict(coverage=1.0, deferrals=0, acc_err=ws_err, acc_mod=tot_mod)
        pol["global-only(tau*)"] = dict(coverage=round(N_acc / N, 4), deferrals=N - N_acc,
                                        acc_err=acc_err, acc_mod=acc_mod)
        pol["beta*-constrained(blocks)"] = dict(coverage=0.0, deferrals=N, acc_err=0, acc_mod=0)
        pol["defer-all"] = pol["beta*-constrained(blocks)"].copy()
        for p in pol.values():
            p["acc_err_per_1000"] = round(1000 * p["acc_err"] / N, 1)
            p["acc_mod_per_1000"] = round(1000 * p["acc_mod"] / N, 1)
        g = pol["global-only(tau*)"]; b = pol["beta*-constrained(blocks)"]
        b["avoided_mod_vs_global"] = g["acc_mod"]
        b["extra_deferrals_vs_global"] = N - g["deferrals"] - 0
        b["deferrals_per_mod_avoided"] = round((N - g["deferrals"]) / g["acc_mod"], 1) if g["acc_mod"] else None
        # DCA proxy: defer = intervention; event = moderate-or-worse erroneous answer
        def nb(deferred_rows):
            TP = sum(sevrule(r) for r in deferred_rows)
            FP = len(deferred_rows) - TP
            return {str(pt): round(TP / N - FP / N * pt / (1 - pt), 4) for pt in PT_GRID}
        dca = {"defer-none": {str(pt): 0.0 for pt in PT_GRID},
               "defer-all": nb(rows), "beta*-constrained(blocks)": nb(rows),
               "global-only(tau*)": nb(rej)}
        return dict(counts=dict(whole_err=ws_err, acc_err=acc_err, rej_err=rej_err,
                                acc_mod=acc_mod, rej_mod=rej_mod), policies=pol, dca_proxy=dca)

    cons_wc = consequences(err_wc, lambda r: modplus_graded(r, err_wc))
    cons_both = consequences(err_both, lambda r: modplus_graded(r, err_both))

    out = dict(run="M6F census certification (Table 9R + decision-consequence proxy)",
               prereg="refine-logs/analysis_plan_v2_20260717.md S4.5/S4.6/S5.2 (frozen)",
               build_id=expect_build, tau_star=0.412661, alpha=ALPHA, beta_star=BETA_STAR,
               N_stream=N, N_accepted=N_acc,
               table9R=t9, decision_at_alpha=decision,
               reliability=dict(kappa_binary_all=round(float(kappa_all), 4),
                                kappa_binary_new832=round(float(kappa_new), 4),
                                raw_agreement_all=round(po, 4), n_unsure_labels=int(n_unsure),
                                intra_rater_repeats=intra),
               harm_worstcase_accepted=dict(distribution=sev_dist,
                                            flagged_ungraded=n_ungraded_flagged),
               consequences_worstcase=cons_wc, consequences_both=cons_both,
               note=("Finite-population exact proportions of the realized stream; CP intervals are "
                     "superpopulation companions under an i.i.d. question-flow reading, labeled as such. "
                     "beta*-constrained gate blocks at the frozen operating point because the certified "
                     "paired bound (robust LCB 0.329) exceeds beta*=0.10; its consequence row therefore "
                     "coincides with defer-all. Label-based decision proxy, not clinical net benefit."))
    with open(artifact("M6F_census_results.json"), "w") as fo:
        json.dump(out, fo, indent=2, ensure_ascii=False)

    print(f"[census] N={N} accepted={N_acc}  build_id OK  unsure-labels={n_unsure}")
    print(f"[reliability] kappa(all 1232)={kappa_all:.4f}  kappa(new 832)={kappa_new:.4f}  "
          f"raw={po:.4f}  intra={intra}")
    print("\nTable 9R — frozen-threshold clinician audit of the complete realized stream")
    for t in t9:
        print(f"  {t['label']:<22} whole {t['whole']:<9} ({t['whole_prop']})   "
              f"accepted {t['accepted']:<9} ({t['accepted_prop']})   mod+ {t['modplus_accepted']}"
              + (f" ({t['modplus_prop']})" if t['modplus_prop'] is not None else ""))
        print(f"  {'':<22} CP95 superpop companion (accepted): {t['cp95_accepted_superpop']}")
    print(f"\n[decision @ alpha={ALPHA}] {decision}")
    print(f"[harm | worst-case accepted] {sev_dist}  (ungraded flagged: {n_ungraded_flagged})")
    g = cons_wc["policies"]["global-only(tau*)"]; b = cons_wc["policies"]["beta*-constrained(blocks)"]
    print(f"[consequences wc] global-only: acc_err/1000={g['acc_err_per_1000']} mod+/1000={g['acc_mod_per_1000']}"
          f" | beta*: avoided mod+={b['avoided_mod_vs_global']} extra deferrals={b['extra_deferrals_vs_global']}"
          f" deferrals-per-mod+-avoided={b['deferrals_per_mod_avoided']}")
    print("[wrote] M6F_census_results.json")


if __name__ == "__main__":
    main()
