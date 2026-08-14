#!/usr/bin/env python
"""M6F — calibration-block consistency check (C01–C20), run BEFORE the formal block is released.

The 20 calibration rows were drawn from the M6-400 audit's BOTH-AGREE rows (10 HALLUCINATED /
10 FAITHFUL under the two clinicians' own prior consensus), so their reference verdicts are
known. This script compares each returned calibration block against that consensus, checks the
build-id (stale/tamper guard), severity completeness on HALLUCINATED verdicts, and inter-doctor
agreement, then writes an aggregate-only JSON (no question/answer text).

Per the frozen protocol (analysis_plan_v2 Section 4): feedback is permitted on calibration rows
only; no performance feedback during the formal block.

Usage: python M6F_check_calibration.py <doctor1_filled.xlsx> <doctor2_filled.xlsx>
"""
import os, sys, csv, json
from openpyxl import load_workbook

from _paths import artifact                                # released layout, docs/LAYOUT.md

HERE = os.path.dirname(os.path.abspath(__file__))          # <package>/code
ROOT = os.path.dirname(HERE)                               # <package>
GREEN_MIN_AGREE = 18   # per doctor, out of 20, vs the M6 consensus


def read_calibration(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["标注表"]
    hdr = [c.value for c in ws[1]]
    col = {k: hdr.index(k) for k in ("build_id", "blind_id", "clinician_verdict", "severity", "notes")}
    rows = {}
    build_ids = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        bid = str(r[col["blind_id"]] or "")
        if not bid.startswith("C"):
            continue
        build_ids.add(str(r[col["build_id"]]))
        rows[bid] = dict(verdict=str(r[col["clinician_verdict"]] or "").strip().upper(),
                         severity=str(r[col["severity"]] or "").strip().lower(),
                         notes=str(r[col["notes"]] or "").strip())
    return rows, build_ids


def main():
    f1, f2 = sys.argv[1], sys.argv[2]
    meta = json.load(open(artifact("M6F_census_meta.json")))
    expect_build = meta["build_id"]

    key = {r["blind_id"]: r for r in csv.DictReader(open(artifact("M6F_census_key.csv")))
           if r["block"] == "calibration"}
    m6v = {}
    for r in csv.DictReader(open(artifact("M6_SR_instrument_annotator1_FILLED.csv"))):
        m6v[r["blind_id"]] = r["clinician_verdict"].strip().upper()
    consensus = {bid: m6v[k["m6_blind_id"]] for bid, k in key.items()}
    assert len(consensus) == 20, len(consensus)

    out = dict(run="M6F calibration check (C01-C20 vs M6-400 both-agree consensus)",
               expect_build_id=expect_build, green_min_agree=GREEN_MIN_AGREE, doctors={})
    verdicts = {}
    for doc, fp in ((1, f1), (2, f2)):
        rows, builds = read_calibration(fp)
        build_ok = builds == {expect_build}
        agree, dis, sev_missing, unfilled = 0, [], [], []
        for bid in sorted(consensus):
            v = rows.get(bid, {}).get("verdict", "")
            if not v:
                unfilled.append(bid); continue
            if v == consensus[bid]:
                agree += 1
            else:
                dis.append(dict(blind_id=bid, given=v, consensus=consensus[bid],
                                note=rows[bid]["notes"][:120]))
            if v == "HALLUCINATED" and not rows[bid]["severity"]:
                sev_missing.append(bid)
        verdicts[doc] = {bid: rows.get(bid, {}).get("verdict", "") for bid in consensus}
        out["doctors"][f"doctor{doc}"] = dict(
            file=os.path.basename(fp), build_id_ok=build_ok, n_filled=20 - len(unfilled),
            n_agree_consensus=agree, disagreements=dis, severity_missing=sev_missing,
            unfilled=unfilled,
            green=(build_ok and not unfilled and not sev_missing and agree >= GREEN_MIN_AGREE))
        print(f"[doctor{doc}] build_id_ok={build_ok} filled={20-len(unfilled)}/20 "
              f"agree_vs_consensus={agree}/20 sev_missing={len(sev_missing)} -> "
              f"{'GREEN' if out['doctors'][f'doctor{doc}']['green'] else 'HOLD'}")
        for d in dis:
            print(f"    disagree {d['blind_id']}: given={d['given']} consensus={d['consensus']}"
                  + (f"  note: {d['note']}" if d['note'] else ""))
    both = [bid for bid in consensus if verdicts[1].get(bid) and verdicts[2].get(bid)]
    inter = sum(1 for bid in both if verdicts[1][bid] == verdicts[2][bid])
    out["inter_doctor_agreement"] = dict(n=len(both), agree=inter)
    out["overall_green"] = all(d["green"] for d in out["doctors"].values())
    print(f"[inter-doctor] {inter}/{len(both)} agree")
    print(f"[overall] {'GREEN — release formal block F0001-F0880' if out['overall_green'] else 'HOLD — see disagreements above'}")
    with open(artifact("M6F_calibration_check.json"), "w") as fo:
        json.dump(out, fo, indent=2, ensure_ascii=False)
    print("[wrote] M6F_calibration_check.json")
    sys.exit(0 if out["overall_green"] else 1)


if __name__ == "__main__":
    main()
