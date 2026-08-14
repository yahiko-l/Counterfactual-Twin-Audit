#!/usr/bin/env python
"""M6F — render the beta* elicitation questionnaire as per-doctor xlsx forms.

Content is the frozen questionnaire (analysis_plan_v2 Section 3 / M6F_make_handoff.ELICITATION_MD),
in the xlsx format the clinicians used in every prior round: a 「问卷说明」 sheet (scenario,
definitions, 12 anchor cases) and a 「作答」 sheet with a dropdown answer cell, an optional
free-value cell, and a reason cell. 填写人 is pre-filled per doctor to avoid mix-ups.

Output (gitignored dir): experiments/M6F_census_handoff/医生{1,2}_容忍度问卷.xlsx
Send FIRST, independently; collect b1, b2; record beta* = min(b1, b2) as a dated amendment
BEFORE sending the census zips.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from _paths import artifact                                # released layout, docs/LAYOUT.md

HERE = os.path.dirname(os.path.abspath(__file__))          # <package>/code
ROOT = os.path.dirname(HERE)                               # <package>
OUTROOT = os.path.join(ROOT, "M6F_census_handoff")
from M6F_make_handoff import ELICITATION_MD  # single source of truth for the frozen text

WRAP = Alignment(wrap_text=True, vertical="top")
FILL = PatternFill("solid", fgColor="FFF2CC")


def build_info_sheet(ws):
    ws.title = "问卷说明"
    ws.column_dimensions["A"].width = 112
    r = 1
    for line in ELICITATION_MD.splitlines():
        c = ws.cell(row=r, column=1, value=line)
        c.alignment = WRAP
        if line.startswith("# "):
            c.value = line[2:]; c.font = Font(bold=True, size=14)
        elif line.startswith("## "):
            c.value = line[3:]; c.font = Font(bold=True, size=12)
        elif line.startswith("> "):
            c.value = line[2:]; c.font = Font(italic=True, color="555555")
        r += 1
    ws.freeze_panes = "A2"


def build_answer_sheet(ws, doc):
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 64
    rows = [
        ("填写人", f"医生{doc}"),
        ("日期", ""),
        ("你的回答 b（每 100 个此类问题中最多可接受的个数）", ""),
        ("如以上候选都不合适，写出你认为更合适的数值", ""),
        ("（可选）一句话说明理由", ""),
        ("确认：填写前未与另一位医生讨论，也未看过本项目任何结果（是/否）", ""),
    ]
    ws.cell(row=1, column=1, value="容忍度问卷 · 作答表").font = Font(bold=True, size=13)
    for i, (k, v) in enumerate(rows, start=3):
        a = ws.cell(row=i, column=1, value=k); a.font = Font(bold=True); a.alignment = WRAP
        b = ws.cell(row=i, column=2, value=v); b.alignment = WRAP
        if not v:
            b.fill = FILL
    dv = DataValidation(type="list", formula1='"10,15,20,25,30,40"', allow_blank=True)
    dv.error = "请从下拉候选中选择；若都不合适请在下一行填写你自己的数值"
    ws.add_data_validation(dv); dv.add("B5")
    dv2 = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
    ws.add_data_validation(dv2); dv2.add("B8")
    ws.cell(row=10, column=1,
            value="填完保存本文件发回任务发起人。判断依据、使用场景与 12 个锚定病例见「问卷说明」工作表。").alignment = WRAP


def main():
    os.makedirs(OUTROOT, exist_ok=True)
    for doc in (1, 2):
        wb = Workbook()
        build_info_sheet(wb.active)
        build_answer_sheet(wb.create_sheet("作答"), doc)
        path = os.path.join(OUTROOT, f"医生{doc}_容忍度问卷.xlsx")
        wb.save(path)
        print(f"wrote {path}")


if __name__ == "__main__":
    main()
