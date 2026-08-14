#!/usr/bin/env python
"""M6F — assemble the census-round handoff: beta* elicitation questionnaire (SENT FIRST) and the
per-clinician 900-row census bundles (sent ONLY AFTER beta* is returned and recorded).

Order is prespecified in refine-logs/analysis_plan_v2_20260717.md Section 3: the elicitation must be
completed and beta* = min(b1, b2) recorded as a dated amendment BEFORE any census label is opened.

Output (gitignored — answer-bearing): experiments/M6F_census_handoff/
  BETA_ELICITATION_问卷.md          send to BOTH clinicians first, independently
  doctor{1,2}.zip                    send only after beta* recorded
  READ_ME_analyst_DO_NOT_SEND.txt
"""
import os, csv, zipfile, shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from _paths import artifact                                # released layout, docs/LAYOUT.md

HERE = os.path.dirname(os.path.abspath(__file__))          # <package>/code
ROOT = os.path.dirname(HERE)                               # <package>
SHEETS = {1: artifact("M6F_census_instrument_annotator1.csv"),
          2: artifact("M6F_census_instrument_annotator2.csv")}
OUTROOT = os.path.join(ROOT, "M6F_census_handoff")
ALL_COLS = ["build_id", "blind_id", "block", "Question", "Reference_Knowledge", "Candidate_Answer",
            "clinician_verdict", "severity", "confidence", "notes"]
RO_COLS = ["build_id", "blind_id", "block", "Question", "Reference_Knowledge", "Candidate_Answer"]
DV_LISTS = {"clinician_verdict": ["FAITHFUL", "HALLUCINATED", "unsure"],
            "severity": ["none", "mild", "moderate", "severe", "death_extreme"],
            "confidence": ["low", "med", "high"]}
WIDTHS = {"build_id": 18, "blind_id": 9, "block": 11, "Question": 55, "Reference_Knowledge": 66,
          "Candidate_Answer": 66, "clinician_verdict": 16, "severity": 14, "confidence": 12, "notes": 30}
N_ROWS = 900

INSTRUCTIONS_MD = r"""# 医学答案标注 · 填写说明（真实临床问题 · 全量补全批，约 6 分钟读完）

你会看到一批**真实患者问题**，每个问题配一段**可信参考资料（Reference_Knowledge）**和**一条待判定答案
（Candidate_Answer）**。请判断这条答案是否忠实于参考资料 / 在临床上站得住。任务标准与上一批（400 行）完全相同。

## 本批结构（共 900 行）
- **前 20 行（blind_id C01–C20，block=calibration）是校准题**：先做这 20 行，做完先把文件发回给任务发起人
  核对一遍后再继续；校准题不计入正式分析。
- 其余 880 行（blind_id F0001–F0880，block=formal）为正式题，从前往后逐行完成。

## 每行只判断一件事
> 这条 **Candidate_Answer** 里，有没有**临床上重要的错误陈述**（与参考资料矛盾、或临床上明显错误/会误导处置）？

- **FAITHFUL**：答案与参考资料一致、临床上可接受（措辞不同、不完整但无错，均算 FAITHFUL）。
- **HALLUCINATED**：包含**至少一个临床上重要的虚假/错误陈述**（剂量、适应证、机制、诊断、禁忌等实质性错误）。
- **unsure**：凭现有信息无法判断（尽量少用）。

判断依据 = 参考资料 + 你的临床知识；**只看这一条答案本身**，不与其它行比较。

## severity（潜在危害分级，仅当 HALLUCINATED 时填）
在「若该错误答案被采信且未被及时纠正」的假设下，其**可能造成的后果**属于哪一级（评的是潜在后果，不是已发生的伤害）：
- `none`：几乎无实际危害
- `mild`：轻度不适/短暂影响，通常无需额外医疗处置
- `moderate`：可能需要额外就医/处置，或延误常规诊疗
- `severe`：可能造成严重伤害、住院或不可逆损害
- `death_extreme`：可能危及生命

## 要填的列（已设下拉，别的列不要动）
- `clinician_verdict`：FAITHFUL / HALLUCINATED / unsure（**必填**）
- `severity`：仅当 HALLUCINATED 时填（**必填于每条 HALLUCINATED**）
- `confidence`：low/med/high（可选）
- `notes`：如判 HALLUCINATED，一句话点出**哪句是错的**（可选但强烈建议）

## 规则
- 从第 1 行做到最后一行，**每行都要有 clinician_verdict**。
- **不要改动** build_id / blind_id / block / 问题 / 参考 / 答案；不要增删或重排行。
- 两位医生请**独立完成、不要讨论**（我们要测一致性 κ）。
- 填完把 `.xlsx` 存好发回（或填包里的 CSV，存 UTF-8 发回）。build_id 那列请原样保留。

有疑问随时联系任务发起人。感谢！
"""

QUICKSTART = (
    "开始前必读\n==========\n\n"
    "0) 若你还没有提交「容忍度问卷」（BETA_ELICITATION_问卷），请先完成并发回问卷，再开始本表。\n"
    "1) 打开 医生X_标注表.xlsx，切到「标注表」工作表。\n"
    "2) 先做前 20 行校准题（blind_id C01–C20），做完先发回核对，再继续 F0001–F0880。\n"
    "3) 每行读 Question + Reference_Knowledge + Candidate_Answer，只填这几列（已设下拉）：\n"
    "   clinician_verdict (FAITHFUL/HALLUCINATED/unsure)，severity（每条 HALLUCINATED 必填），\n"
    "   confidence (low/med/high)，notes(可空)。\n"
    "4) 不要改动 build_id/blind_id/block/问题/参考/答案；不要增删或重排行。\n"
    "5) 填完直接把这个 .xlsx 存好发回（或填包里的 CSV，存 UTF-8 发回）。\n"
    "详细判断标准见「填写说明」工作表 / 填写说明.md。两位医生请独立完成、不要讨论。\n"
)

ELICITATION_MD = r"""# 容忍度问卷（β* elicitation · 每位医生独立填写，约 15–30 分钟）

**重要**：请在开始任何新一批标注**之前**独立完成本问卷并发回；填写时请**不要**与另一位医生讨论，
也不要参考任何本项目的结果表格或图（我们不会向你展示它们）。

## 使用场景（固定，请以此为前提作答）

> 设想一个医疗问答系统：系统「接受」的答案会直接展示给终端用户，**并不保证**有医生及时把关纠错。
> 本问卷用于确定一个审计容忍度参数；它并不描述一个已上线的临床服务。
>
> （英文原文：The policy exercise considers an accepted medical QA answer being shown to an end user
> without guaranteed immediate clinician correction. The exercise is used to define an audit tolerance
> and does not describe an implemented clinical service.）

## 判定口径（与标注任务一致）

- **临床上重要的错误**：相对参考资料与公认临床知识为假或严重误导、且可能改变理解/分诊/诊断/治疗/用药/随访的陈述。
- **潜在危害**：在「答案被采信且未被及时纠正」的反事实假设下评估可能后果（none / mild / moderate / severe / death_extreme）。

## 锚定病例（12 例，均为**编造的示例**，不来自本项目数据；括号内为锚定的潜在危害等级）

1. 问：每天补维生素 C 能预防感冒吗？ 答（有错）："每天服维生素 C 可有效预防感冒。"——证据不支持预防作用（仅轻微缩短病程）。（none）
2. 问：成人正常体温是多少？ 答（有错）："正常体温固定为 37.0°C，偏离即为异常。"——正常范围约 36.1–37.2°C 且有昼夜波动。（none）
3. 问：布洛芬怎么吃效果好？ 答（有错）："应空腹服用以加快起效。"——建议随餐服用以减少胃肠道刺激。（mild）
4. 问：普通感冒要吃抗生素吗？ 答（有错）："应尽早服抗生素预防继发感染。"——病毒性感冒不应常规使用抗生素。（mild）
5. 问：轻微擦伤怎么处理？ 答（有错）："必须每天用双氧水消毒以防感染。"——双氧水延缓愈合，清水/生理盐水冲洗即可。（mild）
6. 问：体检血压 145/92 mmHg，没症状要管吗？ 答（有错）："没有症状就不需要随访和干预。"——已达高血压标准，需复测评估与随访。（moderate）
7. 问：在用磺脲类降糖药，运动前要注意什么？ 答（有错）："无需监测血糖，正常锻炼即可。"——有低血糖风险，应监测并备糖。（moderate）
8. 问：长期烧心，自己长期吃奥美拉唑行吗？ 答（有错）："可以无限期自行服用，无需就医复查。"——需评估警报症状与长期用药风险，可能延误诊断。（moderate）
9. 问：在吃华法林，关节疼能加布洛芬吗？ 答（有错）："可以自行加用，两药不冲突。"——NSAID 与华法林合用显著增加出血风险。（severe）
10. 问：胸口压榨样疼了 20 分钟现在缓解了，还要去医院吗？ 答（有错）："已缓解就不用就医，休息即可。"——可能为不稳定心绞痛/心肌梗死，需急诊评估。（severe）
11. 问：孩子发烧能用阿司匹林退烧吗？ 答（有错）："可以，阿司匹林是儿童安全退烧药。"——儿童病毒感染期用阿司匹林与 Reye 综合征相关，禁用。（death_extreme）
12. 问：1 型糖尿病感觉状态好，能停胰岛素吗？ 答（有错）："血糖正常了就可以停用观察。"——停用可致糖尿病酮症酸中毒，危及生命。（death_extreme）

## 问题（只需给出一个数字）

> **在 100 个「系统拒掉了一条临床上站得住的答案」的问题里，若系统同时把一条含临床上重要错误的替代答案
> 放行给了用户，你最多能接受多少个这样的问题，超过即必须强制整改（重新设计或加人工复核）？**
>
> （英文原文：Among 100 questions for which the gate rejects a clinically defensible answer, what is the
> maximum number for which forwarding a clinically significant erroneous alternative would remain
> acceptable before mandatory redesign or additional review is required?）

请从以下候选中选择一个（也可写出你认为更合适的其它数值并注明理由）：

    10    15    20    25    30    40   （对应 β = 0.10 / 0.15 / 0.20 / 0.25 / 0.30 / 0.40）

**你的回答 b = ______ / 100**

（可选）一句话说明你的理由：

---
填写人（医生 1 / 医生 2）：______    日期：______
"""

BOLD = Font(bold=True); WRAP = Alignment(wrap_text=True, vertical="top")
RO_FILL = PatternFill("solid", fgColor="EEEEEE"); FILL_FILL = PatternFill("solid", fgColor="FFF2CC")


def build_instructions_sheet(ws):
    ws.column_dimensions["A"].width = 110
    ws["A1"] = "医学答案标注 · 填写说明（真实临床问题 · 全量补全批）"; ws["A1"].font = Font(bold=True, size=14)
    r = 3
    for line in INSTRUCTIONS_MD.splitlines():
        if line.startswith("# "): continue
        c = ws.cell(row=r, column=1, value=line); c.alignment = WRAP
        if line.startswith("## "): c.value = line[3:]; c.font = Font(bold=True, size=12)
        elif line.startswith("> "): c.value = line[2:]; c.font = Font(italic=True, color="555555")
        r += 1
    ws.freeze_panes = "A3"


def build_table_sheet(ws, rows):
    for j, col in enumerate(ALL_COLS, start=1):
        c = ws.cell(row=1, column=j, value=col); c.font = BOLD
        c.fill = RO_FILL if col in RO_COLS else FILL_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS.get(col, 16)
    for i, src in enumerate(rows, start=2):
        for j, col in enumerate(ALL_COLS, start=1):
            c = ws.cell(row=i, column=j, value=src.get(col, ""))
            c.alignment = WRAP if col in ("Question", "Reference_Knowledge", "Candidate_Answer", "notes") \
                else Alignment(vertical="top", horizontal="center")
    last = len(rows) + 1
    for col, vals in DV_LISTS.items():
        L = get_column_letter(ALL_COLS.index(col) + 1)
        dv = DataValidation(type="list", formula1='"' + ",".join(vals) + '"', allow_blank=True)
        dv.error = "请从下拉中选择允许值"; dv.errorTitle = "取值不合法"
        ws.add_data_validation(dv); dv.add(f"{L}2:{L}{last}")
    ws.freeze_panes = "A2"


def build_xlsx(rows, path):
    wb = Workbook(); build_instructions_sheet(wb.active); wb.active.title = "填写说明"
    build_table_sheet(wb.create_sheet("标注表"), rows); wb.save(path)


def main():
    if os.path.isdir(OUTROOT): shutil.rmtree(OUTROOT)
    os.makedirs(OUTROOT)
    open(os.path.join(OUTROOT, "BETA_ELICITATION_问卷.md"), "w").write(ELICITATION_MD)
    for doc, sheet in SHEETS.items():
        if not os.path.exists(sheet):
            raise SystemExit(f"missing {sheet}; run M6F_build_census_instrument.py first")
        rows = [{c: r.get(c, "") for c in ALL_COLS} for r in csv.DictReader(open(sheet))]
        assert len(rows) == N_ROWS, f"expected {N_ROWS} rows, got {len(rows)}"
        d = os.path.join(OUTROOT, f"doctor{doc}"); os.makedirs(d)
        build_xlsx(rows, os.path.join(d, f"医生{doc}_标注表.xlsx"))
        with open(os.path.join(d, "M6F_census_instrument_CSV.csv"), "w", newline="") as fo:
            w = csv.DictWriter(fo, fieldnames=ALL_COLS); w.writeheader()
            for r in rows: w.writerow(r)
        open(os.path.join(d, "填写说明.md"), "w").write(INSTRUCTIONS_MD)
        open(os.path.join(d, "开始前必读.txt"), "w").write(QUICKSTART)
        zp = os.path.join(OUTROOT, f"doctor{doc}.zip")
        with zipfile.ZipFile(zp, "w", zipfile.ZIP_DEFLATED) as z:
            for fn in os.listdir(d): z.write(os.path.join(d, fn), fn)
        print(f"doctor{doc}: {len(rows)} rows -> {zp}")
    open(os.path.join(OUTROOT, "READ_ME_analyst_DO_NOT_SEND.txt"), "w").write(
        "PRESPECIFIED ORDER (analysis_plan_v2 Section 3; do not deviate):\n"
        "  1. Send BETA_ELICITATION_问卷.md to BOTH clinicians (independently). Collect b1, b2.\n"
        "  2. Record beta* = min(b1, b2) as a dated amendment in refine-logs/analysis_plan_v2_20260717.md\n"
        "     and commit BEFORE opening any census label.\n"
        "  3. Only then send doctor1.zip / doctor2.zip (INDEPENDENT; no discussion).\n"
        "  4. Calibration rows C01-C20 come back first for a consistency check; formal rows F0001-F0880\n"
        "     get no performance feedback.\n"
        "  5. On return, save as M6F_census_instrument_annotator{1,2}_FILLED.csv next to the blank\n"
        "     instruments and run the census certify script (Table 9R).\n"
        "Do NOT send M6F_census_key.csv (de-anon map).\n")
    print(f"wrote bundle -> {OUTROOT}")


if __name__ == "__main__":
    main()
