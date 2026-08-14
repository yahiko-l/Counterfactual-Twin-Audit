#!/usr/bin/env python
"""M6 — assemble the per-clinician handoff bundle for the fixed-τ S_R global-side audit (400 rows).

Reads the BLINDED instruments M6_SR_instrument_annotator{1,2}.csv (from M6_build_SR_audit_sheet.py) and, per
doctor, writes a self-contained blinded Excel workbook (a 「填写说明」 sheet + a 「标注表」 worksheet with
dropdowns on the label columns), a CSV fallback, a 开始前必读 quickstart, and zips it. The de-anon key
(M6_SR_audit_key.csv) is NEVER placed in a doctor bundle.

Single-answer task (NOT the A/B twin task): each row is ONE candidate answer to a real patient question with a
trusted reference; the clinician judges whether the answer is faithful to the reference / clinically defensible.

Output (gitignored — answer-bearing): experiments/M6_SR_audit_handoff/doctor{1,2}.zip
"""
import os, csv, zipfile, shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from _paths import artifact                                # released layout, docs/LAYOUT.md

HERE = os.path.dirname(os.path.abspath(__file__))          # <package>/code
ROOT = os.path.dirname(HERE)                               # <package>
SHEETS = {1: artifact("M6_SR_instrument_annotator1.csv"),
          2: artifact("M6_SR_instrument_annotator2.csv")}
OUTROOT = os.path.join(ROOT, "M6_SR_audit_handoff")
ALL_COLS = ["build_id", "blind_id", "Question", "Reference_Knowledge", "Candidate_Answer",
            "clinician_verdict", "severity", "confidence", "notes"]
RO_COLS = ["build_id", "blind_id", "Question", "Reference_Knowledge", "Candidate_Answer"]
DV_LISTS = {"clinician_verdict": ["FAITHFUL", "HALLUCINATED", "unsure"],
            "severity": ["none", "mild", "moderate", "severe"],
            "confidence": ["low", "med", "high"]}
WIDTHS = {"build_id": 18, "blind_id": 9, "Question": 55, "Reference_Knowledge": 68,
          "Candidate_Answer": 68, "clinician_verdict": 16, "severity": 12, "confidence": 12, "notes": 30}

INSTRUCTIONS_MD = r"""# 医学答案标注 · 填写说明（真实临床问题 · 全局审计批，约 6 分钟读完）

你会看到一批**真实患者问题**，每个问题配一段**可信参考资料（Reference_Knowledge）**和**一条待判定答案
（Candidate_Answer）**。请判断这条答案是否忠实于参考资料 / 在临床上站得住。

## 每行只判断一件事
> 这条 **Candidate_Answer** 里，有没有**临床上重要的错误陈述**（与参考资料矛盾、或临床上明显错误/会误导处置）？

- **FAITHFUL**：答案与参考资料一致、临床上可接受（措辞不同、不完整但无错，均算 FAITHFUL）。
- **HALLUCINATED**：包含**至少一个临床上重要的虚假/错误陈述**（剂量、适应证、机制、诊断、禁忌等实质性错误）。
- **unsure**：凭现有信息无法判断（尽量少用）。

判断依据 = 参考资料 + 你的临床知识；**只看这一条答案本身**，不与其它行比较。

## 要填的列（已设下拉，别的列不要动）
- `clinician_verdict`：FAITHFUL / HALLUCINATED / unsure（**必填**）
- `severity`：仅当 HALLUCINATED 时填 none/mild/moderate/severe（错误若被采纳对患者的潜在危害）
- `confidence`：low/med/high（可选）
- `notes`：如判 HALLUCINATED，一句话点出**哪句是错的**（可选但强烈建议）

## 规则
- 从第 1 行做到最后一行，**每行都要有 clinician_verdict**。
- **不要改动** build_id / blind_id / 问题 / 参考 / 答案；不要增删或重排行。
- 两位医生请**独立完成、不要讨论**（我们要测一致性 κ）。
- 填完把 `.xlsx` 存好发回（或填包里的 CSV，存 UTF-8 发回）。build_id 那列请原样保留。

有疑问随时联系任务发起人。感谢！
"""

QUICKSTART = (
    "开始前必读\n==========\n\n"
    "1) 打开 医生X_标注表.xlsx，切到「标注表」工作表。\n"
    "2) 从第 1 行往下做，本批共 400 行，全部都要做。\n"
    "3) 每行读 Question + Reference_Knowledge + Candidate_Answer，只填这几列（已设下拉）：\n"
    "   clinician_verdict (FAITHFUL/HALLUCINATED/unsure)，severity（仅 HALLUCINATED 时填），\n"
    "   confidence (low/med/high)，notes(可空)。\n"
    "4) 不要改动 build_id/blind_id/问题/参考/答案；不要增删或重排行。\n"
    "5) 填完直接把这个 .xlsx 存好发回（或填包里的 CSV，存 UTF-8 发回）。\n"
    "详细判断标准见「填写说明」工作表 / 填写说明.md。两位医生请独立完成、不要讨论。\n"
)

BOLD = Font(bold=True); WRAP = Alignment(wrap_text=True, vertical="top")
RO_FILL = PatternFill("solid", fgColor="EEEEEE"); FILL_FILL = PatternFill("solid", fgColor="FFF2CC")


def build_instructions_sheet(ws):
    ws.column_dimensions["A"].width = 110
    ws["A1"] = "医学答案标注 · 填写说明（真实临床问题 · 全局审计批）"; ws["A1"].font = Font(bold=True, size=14)
    r = 3
    for line in INSTRUCTIONS_MD.splitlines():
        if line.startswith("# "): continue
        c = ws.cell(row=r, column=1, value=line); c.alignment = WRAP
        if line.startswith("## "): c.value = line[3:]; c.font = Font(bold=True, size=12)
        elif line.startswith("> "): c.value = line[2:]; c.font = Font(italic=True, color="555555")
        r += 1
    ws.freeze_panes = "A3"


def build_table_sheet(ws, rows):
    for j, col in enumerate(ALL_COLS, start=1):
        c = ws.cell(row=1, column=j, value=col); c.font = BOLD
        c.fill = RO_FILL if col in RO_COLS else FILL_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS.get(col, 16)
    for i, src in enumerate(rows, start=2):
        for j, col in enumerate(ALL_COLS, start=1):
            c = ws.cell(row=i, column=j, value=src.get(col, ""))
            c.alignment = WRAP if col in ("Question", "Reference_Knowledge", "Candidate_Answer", "notes") \
                else Alignment(vertical="top", horizontal="center")
    last = len(rows) + 1
    for col, vals in DV_LISTS.items():
        L = get_column_letter(ALL_COLS.index(col) + 1)
        dv = DataValidation(type="list", formula1='"' + ",".join(vals) + '"', allow_blank=True)
        dv.error = "请从下拉中选择允许值"; dv.errorTitle = "取值不合法"
        ws.add_data_validation(dv); dv.add(f"{L}2:{L}{last}")
    ws.freeze_panes = "A2"


def build_xlsx(rows, path):
    wb = Workbook(); build_instructions_sheet(wb.active); wb.active.title = "填写说明"
    build_table_sheet(wb.create_sheet("标注表"), rows); wb.save(path)


def main():
    if os.path.isdir(OUTROOT): shutil.rmtree(OUTROOT)
    os.makedirs(OUTROOT)
    for doc, sheet in SHEETS.items():
        if not os.path.exists(sheet):
            raise SystemExit(f"missing {sheet}; run M6_build_SR_audit_sheet.py first")
        rows = [{c: r.get(c, "") for c in ALL_COLS} for r in csv.DictReader(open(sheet))]
        assert len(rows) == 400, f"expected 400 rows, got {len(rows)} (did you build --full?)"
        d = os.path.join(OUTROOT, f"doctor{doc}"); os.makedirs(d)
        build_xlsx(rows, os.path.join(d, f"医生{doc}_标注表.xlsx"))
        with open(os.path.join(d, "M6_SR_instrument_CSV.csv"), "w", newline="") as fo:
            w = csv.DictWriter(fo, fieldnames=ALL_COLS); w.writeheader()
            for r in rows: w.writerow(r)
        open(os.path.join(d, "填写说明.md"), "w").write(INSTRUCTIONS_MD)
        open(os.path.join(d, "开始前必读.txt"), "w").write(QUICKSTART)
        zp = os.path.join(OUTROOT, f"doctor{doc}.zip")
        with zipfile.ZipFile(zp, "w", zipfile.ZIP_DEFLATED) as z:
            for fn in os.listdir(d): z.write(os.path.join(d, fn), fn)
        print(f"doctor{doc}: {len(rows)} rows -> {zp}")
    open(os.path.join(OUTROOT, "READ_ME_analyst_DO_NOT_SEND.txt"), "w").write(
        "Send doctor1.zip to clinician 1 and doctor2.zip to clinician 2 (INDEPENDENT).\n"
        "Do NOT send M6_SR_audit_key.csv (de-anon map). On return, save the two filled sheets as\n"
        "M6_SR_instrument_annotator{1,2}_FILLED.csv and run: python M6_certify_SR_audit.py\n")
    print(f"wrote bundle -> {OUTROOT}")


if __name__ == "__main__":
    main()
