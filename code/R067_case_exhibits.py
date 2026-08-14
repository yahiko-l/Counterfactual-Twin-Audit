#!/usr/bin/env python
"""R067 — worked paired-failure case exhibits (reviewer Minor 6), selected by the PRESPECIFIED
rule (analysis_plan_v2 S5.5, amended 2026-07-18: severity-descending then blind-HMAC, because the
census produced more severe cases than the exhibit budget). NO content-based selection.

Source: the census worst-case-flagged ACCEPTED rows (either clinician labels a clinically
significant error) at the frozen tau*, with the maximum available potential-harm grade. Selection:
all death/extreme + severe first, then fill to 8 (2 main-text + 6 supplementary) by descending
severity, ties broken by the lowest HMAC of the blind_id under a fixed key. This script writes:
  * a COMMITTABLE de-identified manifest (blind_id, source, harm grade, selection rank, HMAC) —
    NO question/answer text, so it is safe to version;
  * a LOCAL (gitignored) analyst exhibit sheet with the actual question/reference/candidate and
    the three scores, for the authors to anonymize into the paper by hand.

Output: R067_case_exhibits_manifest.json (committable) + R067_case_exhibits_LOCAL.json (gitignored).
"""
import os, sys, csv, json, hmac, hashlib

from _paths import artifact                                # released layout, docs/LAYOUT.md

HERE = os.path.dirname(os.path.abspath(__file__))          # <package>/code
ROOT = os.path.dirname(HERE)                               # <package>
HMAC_KEY = b"psig-exhibit-2026"   # fixed, published in the manifest note (selection is reproducible, not secret)
SEV_RANK = {"death_extreme": 4, "severe": 3, "moderate": 2, "mild": 1, "": 0, "none": 0}
N_MAIN, N_SUPP = 2, 6


def read_filled(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True)["标注表"]
    hdr = [c.value for c in ws[1]]
    col = {k: hdr.index(k) for k in ("blind_id", "clinician_verdict", "severity")}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        bid = str(r[col["blind_id"]] or "")
        if bid:
            out[bid] = (str(r[col["clinician_verdict"]] or "").strip().upper(),
                        str(r[col["severity"]] or "").strip().lower())
    return out


def main():
    f1, f2 = sys.argv[1], sys.argv[2]
    d1, d2 = read_filled(f1), read_filled(f2)
    key = {r["blind_id"]: r for r in csv.DictReader(open(artifact("M6F_census_key.csv")))}
    # M6-400 rows too (accepted audit); merge their verdicts
    m6 = {}
    for a, dd in ((1, {}), (2, {})):
        for r in csv.DictReader(open(artifact(f"M6_SR_instrument_annotator{a}_FILLED.csv"))):
            dd[r["blind_id"]] = (r["clinician_verdict"].strip().upper(), (r.get("severity") or "").strip().lower())
        m6[a] = dd
    m6key = {r["blind_id"]: r for r in csv.DictReader(open(artifact("M6_SR_audit_key.csv")))
             if r["blind_id"] != ""}

    # unify accepted rows with a worst-case error flag + max harm grade
    cands = []
    def consider(bid, src, accepted, v1, s1, v2, s2, origin):
        if accepted != 1:
            return
        wc_err = v1 in ("HALLUCINATED", "UNSURE") or v2 in ("HALLUCINATED", "UNSURE")
        if not wc_err:
            return
        sev = max(SEV_RANK.get(s1, 0), SEV_RANK.get(s2, 0))
        h = hmac.new(HMAC_KEY, bid.encode(), hashlib.sha256).hexdigest()
        cands.append(dict(blind_id=bid, source=src, origin=origin, harm_rank=sev,
                          harm=[k for k, v in SEV_RANK.items() if v == sev and k][0] if sev else "ungraded",
                          hmac=h))

    for bid, k in key.items():
        if k["block"] != "formal" or k["kind"] != "main":   # exclude the 48 hidden repeats (duplicate labelings)
            continue
        if bid in d1 and bid in d2:
            consider(bid, k["source"], int(k["accepted"]), d1[bid][0], d1[bid][1], d2[bid][0], d2[bid][1], "census")
    for bid, k in m6key.items():
        consider(bid, k["source"], 1, m6[1][bid][0], m6[1][bid][1], m6[2][bid][0], m6[2][bid][1], "audit400")

    cands.sort(key=lambda c: (-c["harm_rank"], c["hmac"]))
    chosen = cands[:N_MAIN + N_SUPP]
    for i, c in enumerate(chosen):
        c["placement"] = "main" if i < N_MAIN else "supplementary"
        c["rank"] = i + 1

    manifest = dict(run="R067 case-exhibit selection (prespecified severity-desc + blind-HMAC)",
                    rule="all death/extreme+severe first, then descending harm, ties by lowest HMAC(blind_id); no content-based selection",
                    hmac_key=HMAC_KEY.decode(), n_worstcase_accepted=len(cands),
                    harm_histogram={h: sum(c["harm"] == h for c in cands)
                                    for h in ("death_extreme", "severe", "moderate", "mild", "ungraded")},
                    exhibits=[{k: c[k] for k in ("rank", "placement", "blind_id", "source", "harm", "harm_rank")}
                              for c in chosen])
    json.dump(manifest, open(artifact("R067_case_exhibits_manifest.json"), "w"), indent=2)

    # local analyst sheet with real text (gitignored)
    txt = {}
    import glob
    for fp in glob.glob(artifact("natural_real_*.jsonl")):
        for l in open(fp):
            r = json.loads(l)
            if r.get("model_answer") and r.get("answerer"):
                txt[(r["source"], str(r["qid"]), r["answerer"])] = r["model_answer"]
    qk = {}
    for l in open(artifact("real_sources_questions.jsonl")):
        r = json.loads(l); qk[(r["source"], str(r["qid"]))] = (r.get("Q", ""), r.get("K", ""))
    local = []
    allkey = {**{b: key[b] for b in key}, **{b: dict(m6key[b], block="audit400") for b in m6key}}
    for c in chosen:
        k = allkey.get(c["blind_id"], {})
        src, qid, ans = k.get("source", ""), k.get("qid", ""), k.get("answerer", "")
        Q, Kref = qk.get((src, str(qid)), ("", ""))
        local.append(dict(rank=c["rank"], placement=c["placement"], harm=c["harm"], source=src,
                          question=Q, reference=Kref, candidate=txt.get((src, str(qid), ans), ""),
                          note="anonymize before use; drop identifiers; verify harm rationale"))
    json.dump(local, open(artifact("R067_case_exhibits_LOCAL.json"), "w"),
              indent=2, ensure_ascii=False)

    print(f"[cands] {len(cands)} worst-case accepted; harm hist = {manifest['harm_histogram']}")
    for c in chosen:
        print(f"  #{c['rank']} [{c['placement']:<13}] {c['blind_id']:<7} {c['source']:<14} harm={c['harm']}")
    print("[wrote] R067_case_exhibits_manifest.json (committable) + R067_case_exhibits_LOCAL.json (gitignored)")


if __name__ == "__main__":
    main()
